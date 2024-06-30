# 引入包
import openpyxl
from openpyxl import load_workbook
import requests
import json
import time
import random9


# 请求地址
Host = 'https://api.youpin898.com/api'

# api接口
GetCsGoPagedList = '/homepage/v2/es/commodity/GetCsGoPagedList'  # 单个参数
list = '/api/homepage/Search/new/Search'




def market(number):
    r = []
    r = requests.post(url=Host + GetCsGoPagedList, json={"templateId": number, "userId": 3250568, })
    data = r.json()

    # 悠悠写入execl
    file_path = 'UU.xlsx'
    sheet_name = 'Sheet1'
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # market(10262)

    sheetTemplateId = data['Data']['CommodityList'][0]['TemplateId']
    sheetCommodityName = data['Data']['CommodityList'][0]['CommodityName']
    sheetPrice = data['Data']['CommodityList'][0]['Price']
    sheetTypeName = data['Data']['TemplateInfo']['TypeName']
    sheetExterior = data['Data']['TemplateInfo']['Exterior']
    sheetRarity = data['Data']['TemplateInfo']['Rarity']
    data = [sheetTemplateId, sheetCommodityName, sheetPrice, sheetTypeName, sheetExterior, sheetRarity]
    sheet.append(data)
    workbook.save(file_path)


counter = 44

while True:
    try:
        market(counter)
        time.sleep(random.randint(8, 10))
        counter += 1
        print(counter)
    except TypeError:
        print('第' + str(counter + 1) + '次参数错误')
        counter += 1

# TemplateId= data['Data']['CommodityList'][0]['TemplateId']
# CommodityName=data['Data']['CommodityList'][0]['CommodityName']
# Price=data['Data']['CommodityList'][0]['Price']

# 悠悠写入execl
# file_path = 'UU.xlsx'
# sheet_name = 'Sheet1'
# workbook = load_workbook(file_path)
# sheet = workbook[sheet_name]
# sheet['A1'] ='templateId'
# sheet['B1'] ='CommodityName'
# sheet['C1'] ='Price'
# sheet['A2'] = TemplateId
# sheet['B2'] =CommodityName
# sheet['C2'] =Price
#
# workbook.save(file_path)
